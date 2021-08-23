import os
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
import xlsxwriter

chrome_options = Options()
chrome_options.add_argument('--headless')
# driver = webdriver.Chrome(chrome_options=chrome_options) 
driver = webdriver.Chrome()
# time.sleep(10)
driver.get('http://sydwbm.rsj.nantong.gov.cn')
select = driver.find_element_by_name('jobAreaList')
jobAreaLists = select.find_elements_by_tag_name('option')
area_list = []
new_job = []
for jobAreaList in jobAreaLists:
    if '请选择' not in jobAreaList.text:
        # jobAreaLists.remove(jobAreaList)
        area_list.append(jobAreaList.text)
for i in range(1,len(area_list)+1):
    Select(select).select_by_index(i)
    # print('start')
    driver.find_element_by_class_name('lgbtn').submit()
    jobs = driver.find_elements_by_tag_name('tr')
    for job in jobs:
        new_job.append(job.text)
    select = driver.find_element_by_name('jobAreaList')
jobs = []
for job in new_job:
    if '部门名称 职位名称' not in job:
        jobs.append(job)
# print(jobs)
driver.quit()

file = 'e:\报名信息.xlsx'
if os.path.exists(file):
        os.remove(file)
workbook = xlsxwriter.Workbook(file)   #创建一个Excel文件
worksheet = workbook.add_worksheet('所有报名信息')  #创建一个sheet
title = [U'部门名称',U'部门代码',U'职位名称',U'职位代码',U'信息审核通过',U'未审核',U'报名成功','总人数']
worksheet.write_row('A1',title)

tag_dept = ['海门区中医院','海门区四甲镇综合服务中心','通州区规划服务中心','通州区劳动就业管理中心','通州区应急管理服务中心','启东市第二人民医院','启东市医疗保险基金管理中心','南通市崇川区区域治理现代化指挥中心','海安市市域治理现代化指挥中心','海安市殡仪馆','海安综合服务中心','西场综合服务中心','南莫镇综合服务中心','李堡镇综合服务中心','如东县公共资源交易中心','如东县东安新闸管理所','如东县掘港殡仪馆','如东县岔河镇中心卫生院']
row = 1
for m in range(0,len(jobs)):
    for n in tag_dept:
        if n in jobs[m] :
            dept_name = jobs[m].split(' ')[0]
            dept_code = jobs[m].split(' ')[1]
            job_name  = jobs[m].split(' ')[2]
            job_code = jobs[m].split(' ')[3]
            checked = jobs[m].split(' ')[4]
            check = jobs[m].split(' ')[5]
            success = jobs[m].split(' ')[6]
            total = int(checked)+int(check)+int(success)
            worksheet.write(row,0,dept_name)
            worksheet.write(row,1,dept_code)
            worksheet.write(row,2,job_name)
            worksheet.write(row,3,job_code)
            worksheet.write(row,4,checked)
            worksheet.write(row,5,check)
            worksheet.write(row,6,success)
            worksheet.write(row,7,total)
            row = row+1
workbook.close()

# file = 'e:\报名信息.xlsx'
# index = [5,6,7,11,12,13,17,18,19,20,21,22,23,25,28,29,31,34,37]
rows = [5,6,7,11,12,13,17,18,19,20,21,22,23,25,28,29,31,34,37]
new_rows = []
for m in rows:
    new_rows.append(int(m)- rows.index(m))
# print(new_rows)

xlxs = load_workbook(file)
ws = xlxs.active
for i in new_rows:
    # delete_cols 删除列
    # delete_rows 删除行
    ws.delete_rows(idx = i ,amount = 1)
xlxs.save(file)